from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from utils import extract_1688_offer_id, now_timestamp


TEMPLATE_SHEET_NAME = "模板"
HEADER_ROW = 3
DATA_ROW = 4


def build_excel_sku(source_url: str, fallback_prefix: str = "SKU") -> str:
    offer_id = extract_1688_offer_id(source_url)
    if offer_id:
        return f"1688_{offer_id}"
    return f"{fallback_prefix}_{now_timestamp()}"


def _load_workbook_from_bytes(template_bytes: bytes):
    return load_workbook(BytesIO(template_bytes))


def get_template_headers(template_bytes: bytes) -> list[str]:
    wb = _load_workbook_from_bytes(template_bytes)
    ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.active
    return [str(ws.cell(HEADER_ROW, col).value or "").strip() for col in range(1, ws.max_column + 1)]


def get_data_options(template_bytes: bytes, data_column_index: int) -> list[str]:
    wb = _load_workbook_from_bytes(template_bytes)
    if "Data" not in wb.sheetnames:
        return []
    ws = wb["Data"]
    values: list[str] = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, data_column_index).value
        if value not in (None, ""):
            text = str(value)
            if text not in values:
                values.append(text)
    return values


def _columns_from_sqref(sqref: Any) -> set[int]:
    columns: set[int] = set()
    for range_part in str(sqref).split():
        min_col, _, max_col, _ = range_boundaries(range_part)
        columns.update(range(min_col, max_col + 1))
    return columns


def _options_from_formula(wb: Any, formula: str | None) -> list[str]:
    if not formula or "!" not in formula:
        return []
    sheet_name, cell_range = formula.split("!", 1)
    sheet_name = sheet_name.strip("'")
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range.replace("$", ""))
    values: list[str] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value not in (None, ""):
                text = str(cell.value)
                if text not in values:
                    values.append(text)
    return values


def get_template_fields(template_bytes: bytes) -> list[dict[str, Any]]:
    wb = _load_workbook_from_bytes(template_bytes)
    ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.active

    options_by_col: dict[int, list[str]] = {}
    for validation in ws.data_validations.dataValidation:
        if validation.type != "list":
            continue
        options = _options_from_formula(wb, validation.formula1)
        if not options:
            continue
        for col in _columns_from_sqref(validation.sqref):
            options_by_col[col] = options

    fields: list[dict[str, Any]] = []
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(HEADER_ROW, col).value or "").strip()
        if not header:
            continue
        note = str(ws.cell(2, col).value or "").strip()
        group = str(ws.cell(1, col).value or "").strip()
        fields.append(
            {
                "column": col,
                "letter": get_column_letter(col),
                "header": header,
                "required": "*" in header,
                "group": group,
                "note": note,
                "options": options_by_col.get(col, []),
            }
        )
    return fields


def fill_official_excel_template(
    template_bytes: bytes,
    values_by_header: dict[str, Any],
    output_path: str | Path,
) -> Path:
    return fill_official_excel_template_rows(template_bytes, [values_by_header], output_path)


def _copy_template_row(ws: Any, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)
        if source.border:
            target.border = copy(source.border)


def fill_official_excel_template_rows(
    template_bytes: bytes,
    rows_by_header: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    wb = _load_workbook_from_bytes(template_bytes)
    ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.active

    headers_by_col = {
        col: str(ws.cell(HEADER_ROW, col).value or "").strip()
        for col in range(1, ws.max_column + 1)
    }

    for index, values_by_header in enumerate(rows_by_header):
        row_number = DATA_ROW + index
        _copy_template_row(ws, DATA_ROW, row_number)
        for col, header in headers_by_col.items():
            cell = ws.cell(row_number, col)
            cell.value = values_by_header.get(header)

    target = Path(output_path)
    if not target.is_absolute():
        target = Path.cwd() / target
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def build_hoodie_template_values(
    *,
    source_url: str,
    ru_title: str,
    en_title: str,
    ru_description: str,
    en_description: str,
    image_urls: list[str],
    price: str,
    discount_price: str,
    inventory: int,
    shipping_lead_time: int,
    weight: str,
    package_length: int,
    package_width: int,
    package_height: int,
    shipping_template_name: str,
    sale_method: str,
    origin: str,
    hoodie_type: str,
    material: str,
    gender: str,
    style: str,
    item_type: str,
    color: str,
    size: str,
) -> dict[str, Any]:
    sku = build_excel_sku(source_url)
    values: dict[str, Any] = {
        "型号（卖家SPU ID）*": sku,
        "SKU ID/条形码（卖家SKU ID）*": sku,
        "产品名称 (俄语)*": ru_title,
        "产品名称 (英语)": en_title,
        "描述 (俄语)*": ru_description,
        "描述 (英语)": en_description,
        "销售方式*": sale_method,
        "价格, CNY*": price,
        "库存": inventory,
        "订单处理时间（天）*": shipping_lead_time,
        "包装重量（公斤）*": weight,
        "包装长度（厘米）*": package_length,
        "包装宽度（厘米）*": package_width,
        "包装高度（厘米）*": package_height,
        "运输模板*": shipping_template_name,
        "Origin": origin,
        "Type": hoodie_type,
        "Material": material,
        "Gender": gender,
        "Style": style,
        "Item Type": item_type,
        "Color": color,
        "Size": size,
    }
    if discount_price.strip():
        values["折扣价, CNY"] = discount_price

    for index, image_url in enumerate(image_urls[:6], start=1):
        values["主图*" if index == 1 else f"图片 #{index}"] = image_url
    return values


def build_common_template_values(
    *,
    fields: list[dict[str, Any]],
    source_url: str,
    ru_title: str,
    en_title: str,
    ru_description: str,
    en_description: str,
    image_urls: list[str],
    price: str,
    discount_price: str,
    inventory: int,
    shipping_lead_time: int,
    weight: str,
    package_length: int,
    package_width: int,
    package_height: int,
) -> dict[str, Any]:
    sku = build_excel_sku(source_url)
    values: dict[str, Any] = {}
    image_index = 0

    for field in fields:
        header = field["header"]
        normalized = header.replace(" ", "").lower()

        if "spuid" in normalized or "spu" in normalized or "型号" in header:
            values[header] = sku
        elif "sku" in normalized or "条形码" in header:
            values[header] = sku
        elif "产品名称" in header and "俄语" in header:
            values[header] = ru_title
        elif "产品名称" in header and "英语" in header:
            values[header] = en_title
        elif "描述" in header and "俄语" in header:
            values[header] = ru_description
        elif "描述" in header and "英语" in header:
            values[header] = en_description
        elif header == "主图*" or header.startswith("图片 #"):
            if image_index < len(image_urls):
                values[header] = image_urls[image_index]
            image_index += 1
        elif "价格" in header and "折扣" not in header:
            values[header] = price
        elif "折扣价" in header:
            values[header] = discount_price
        elif "库存" in header:
            values[header] = inventory
        elif "订单处理时间" in header:
            values[header] = shipping_lead_time
        elif "包装重量" in header:
            values[header] = weight
        elif "包装长度" in header:
            values[header] = package_length
        elif "包装宽度" in header:
            values[header] = package_width
        elif "包装高度" in header:
            values[header] = package_height

    return {key: value for key, value in values.items() if value not in (None, "")}
